from docx import Document as dx
import pandas as pd
import re
import os
import ast  # safer than eval for literal evaluation
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from langchain_chroma import Chroma
from model.bedrock import bedrock_embedding,ask_claude
from prompts.gap_finder_prompt import build_gap_prompt,build_table_prompt
from langchain.schema.document import Document

# Parse the "Embedding" column
def parse_embedding(embedding_str):
    try:
        return np.array(ast.literal_eval(embedding_str))  # Convert to numpy array
    except Exception as e:
        print(f"Error parsing embedding: {embedding_str}, error: {e}")
        return None


def split_docx_by_bold_titles(file_path):
    # Open the .docx file
    doc = dx(file_path)
    
    # Initialize storage for titles and paragraphs as a list of lists
    content = []
    current_title = None
    index = 1
    main_title = None

    # Iterate through all elements in the document
    for paragraph in doc.paragraphs:
        if paragraph.style.name == 'List Paragraph':
            main_title = str(index) + '. ' + paragraph.text.strip()
            index += 1
            # Ensure a new entry for the main_title with an empty list
            content.append([main_title, "", []])
            current_title = None  # Reset current_title when a new main_title is encountered
        # Check if the paragraph is bold (at least one run is bold)
        elif paragraph.style.name != 'List Paragraph' and any(run.bold for run in paragraph.runs):
            current_title = paragraph.text.strip()
            # Add a new sublist for the current_title under the current main_title
            content.append([main_title, current_title, []])
        elif main_title:
            # If no current_title, add paragraphs directly under "No title"
            text = paragraph.text.strip()
            if text:  # Ignore empty paragraphs
                if current_title:  # Append to the current_title
                    content[-1][2].append(text)
                else:  # Append to "No title"
                    content[-1][2].append(text)

    return content

def create_vectorstore(db_directory,split_content):
    if not os.path.isdir(db_directory):
        to_embed_docs=[]
        doc_num=1
        for row in split_content:
            
            # to_embed.append(f"Title: {title}\n"+joined_par)
            if row[2]:
                joined_par='\n'.join(row[2])
                if row[1]:
                    content=f"Title: {row[0]}\n SubTitle: {row[1]}\n"+joined_par

                else:
                    content=f"Title: {row[0]}\n"+joined_par
                print(content)
                to_embed_docs.append(Document(page_content =content,
                                    metadata = {
                                                'id':doc_num,
                                                'title':row[0],
                                                'sub_title':row[1]}))
            doc_num+=1
        # print(to_embed[0])
        os.makedirs(db_directory)
        vectorstore=Chroma(
            collection_name="collection_document",
            embedding_function=bedrock_embeddings,
            persist_directory=db_directory,  # Where to save data locally, remove if not neccesary
        )
        vectorstore.add_documents(to_embed_docs)
    else: 
        vectorstore=Chroma(
            collection_name="collection_document",
            embedding_function=bedrock_embeddings,
            persist_directory=db_directory,  # Where to save data locally, remove if not neccesary
        )
    
    return vectorstore 

def extract_table_from_text(Article,Original_article,text):
    """
    Extracts a table from a given string and converts it into a Pandas DataFrame.

    Args:
        text (str): The input string containing the table.

    Returns:
        pd.DataFrame: A DataFrame representing the table.
    """
    # Split lines and filter out separators
    lines = text.split("\n")
    rows = [line.strip("|").split("|") for line in lines if "|" in line and not re.match(r"^\|-+", line)]

    # Check if we have at least headers and one data row
    if len(rows) < 2:
        # raise ValueError("The provided string does not contain a valid table.")
        data=[[f"Rz.{Article}", Original_article ,'','','','The provided string does not contain a valid table.']]

    # Extract headers and data rows

    data = [[f"Rz.{Article}"]+[Original_article]+ [cell.replace('<br>','\n').strip() for cell in row] for row in rows[1:]]
    
    for row in data:
        if row[3] == "Yes":
            row[3] = "✓"
        elif row[3] == "No":
            row[3] = "×"

    # Create the DataFrame
    return data

def Write_to_excel(df_table):
    # Write DataFrame to an Excel file
    with pd.ExcelWriter('Results/Report_2017_formated_selected.xlsx', engine='openpyxl') as writer:
        df_table.to_excel(writer, index=False, sheet_name="Sheet1")

        # Access the workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Sheet1']

        # Define styles for header and rows
        header_fill = PatternFill(start_color="434fc3", end_color="434fc3", fill_type="solid")  # Blue for header
        even_row_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")    # Gray for even rows
        header_font = Font(bold=True, color="FFFFFF", size=12)  # Bold white font for header
        green_font = Font(color="008000", bold=True, size=12) 
        red_font = Font(color="fe4017", bold=True, size=14) 
        centered_alignment = Alignment(horizontal='center', vertical='center')
        wrapping_alignment = Alignment(wrap_text=True, vertical='top') 
        bottom_border = Border(bottom=Side(style="thin"))  # Thin bottom border
        # Apply header styles
        for cell in worksheet[1]:  # Header row
            cell.fill = header_fill
            cell.font = header_font

        # Apply even row styles
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            worksheet.cell(row=row_idx, column=3).alignment = wrapping_alignment
            worksheet.cell(row=row_idx, column=5).alignment = wrapping_alignment
            worksheet.cell(row=row_idx, column=6).alignment = wrapping_alignment
        
            if row_idx % 2 == 1:  # Even row
                for cell in row[2:]:
                    cell.fill = even_row_fill
            
            covered_cell = row[3]
            covered_cell.alignment=centered_alignment
            if covered_cell.value == "✓":
                covered_cell.font = green_font                
            elif covered_cell.value == "×":
                covered_cell.font = red_font


        # Set column widths
        column_widths = {
            'A': 10,  # Width for "Article"
            'B': 40, # Width for "Content"
            'C': 30,  # Width for "Requirements"
            'D': 10,  # Width for "Covered"
            'E': 30,  # Width for "Reference in Document"
            'F': 60   # Width for "Comment"
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width 


        # Merge cells in the "Article" column if they have the same value
        current_article = None
        start_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            article_cell = row[0]  # Column A
            if article_cell.value != current_article:
                if start_row is not None:
                    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=row_idx - 1, end_column=1)
                    worksheet.merge_cells(start_row=start_row, start_column=2, end_row=row_idx - 1, end_column=2)
                    worksheet.cell(row=start_row, column=1).alignment = centered_alignment
                    worksheet.cell(row=start_row, column=2).alignment = wrapping_alignment
                    # Apply a bottom border to all cells in the last row of the merge
                    for col_idx in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_idx - 1, column=col_idx).border = bottom_border
                current_article = article_cell.value
                start_row = row_idx
        # Merge the last group if applicable
        if start_row is not None:
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=worksheet.max_row, end_column=1)
            worksheet.merge_cells(start_row=start_row, start_column=2, end_row=worksheet.max_row, end_column=2)
            worksheet.cell(row=start_row, column=1).alignment = centered_alignment 
            worksheet.cell(row=start_row, column=2).alignment = wrapping_alignment
            # Apply a bottom border to all cells in the last row of the merge
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.cell(row=worksheet.max_row, column=col_idx).border = bottom_border

        worksheet.column_dimensions.group('B', 'B', hidden=True) 

file_path = "Data/Document/Concept_Risk and Governance_EN.docx"  
split_content= split_docx_by_bold_titles(file_path)

bedrock_embeddings=bedrock_embedding()
db_directory_document= f"vectorestores/chroma_db_document"
vectorstore=create_vectorstore(db_directory_document,split_content)



df_2017=pd.read_excel('Data/Finma_EN/splitted/finma2017.xlsx') 
df_2017['Embedding'] = df_2017['Embedding'].apply(parse_embedding)
# print(type(list(df_2017['Embedding'][0])))
responsedf= pd.read_excel('Results/GapAnalyzer_Finma_Reporting.xlsx') 
results_df_2017 = pd.DataFrame(columns=['Margin','Finma2017', "Reports","Table"])

headers = ['Article','Article Content','Requirement','Covered', 'Reference in Document', 'Comment']
all_rows=[]
for index,response in enumerate(responsedf['Reports']):
    table_prompt= build_table_prompt(response) 
    table_response=ask_claude(table_prompt,0.1)
    margin=responsedf['Margin'][index]
    full_text=responsedf['Finma2017'][index]
    cur_table_data= extract_table_from_text(margin,full_text,table_response)
    all_rows=all_rows+cur_table_data

subset_df = df_2017.iloc[9:66]
headers = ['Article','Article Content','Requirement','Covered', 'Reference in Document', 'Comment']
all_rows=[]
for i, row in subset_df.iterrows():
    full_text=''
    title = str(row.get('Title', None))
    sub_title = str(row.get('SubTitle', None))
    sub_subtitle = str(row.get('Sub_Subtitle', None))
    margin = str(row.get('Margin', None))
    text = row.get('Text', None)
    embedded_item = row.get('Embedding', None)

    if text !='Abrogated':
        results = vectorstore.similarity_search_by_vector(
            embedding=embedded_item , k=4
        )

        # response=ask_claude(gap_prompt,0.6)
        if title and title!='None' and title!='nan':
            full_text=title
        if sub_title and sub_title!='None' and sub_title!='nan':
            full_text=full_text+'\n'+sub_title
        if sub_subtitle and sub_subtitle!='None' and sub_subtitle!='nan':
            full_text=full_text+'\n'+sub_subtitle
        if full_text:
           full_text=full_text +'\n'+text
        else:
            full_text=text

        gap_prompt=build_gap_prompt(results,full_text)
        
        response=ask_claude(gap_prompt,0.6)

        table_prompt= build_table_prompt(response)   
        # print(table_prompt) 
        table_response=ask_claude(table_prompt,0.1) 

        new_row = [{ 
                    "Margin" : margin,
                    "Finma2017" : full_text,
                    "Reports": response,
                    "Table": table_response
                }]
        results_df_2017 = pd.concat([results_df_2017, pd.DataFrame(new_row)], ignore_index=True)

        
        
        cur_table_data= extract_table_from_text(margin,full_text,table_response)
        all_rows=all_rows+cur_table_data
        # print(response)
results_df_2017.to_excel('Results/Report_2017_0.6.xlsx', index=False) 
df_table= pd.DataFrame(all_rows, columns=headers)  
Write_to_excel(df_table)



   








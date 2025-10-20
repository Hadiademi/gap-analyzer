import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook

def write_to_excel(df_table):
    # Create an in-memory BytesIO buffer
    output = io.BytesIO()
    
    # Write DataFrame to an Excel file in memory
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_table.to_excel(writer, index=False, sheet_name="Sheet1")

        workbook = writer.book
        worksheet = workbook["Sheet1"]

        # Define styles
        header_fill = PatternFill(start_color="434fc3", end_color="434fc3", fill_type="solid")  
        even_row_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        green_font = Font(color="008000", bold=True, size=12)
        red_font = Font(color="fe4017", bold=True, size=14)
        centered_alignment = Alignment(horizontal="center", vertical="center")
        wrapping_alignment = Alignment(wrap_text=True, vertical="top")
        center_and_wrap=  Alignment(wrap_text=True,horizontal="center", vertical="center")
        bottom_border = Border(bottom=Side(style="thin"))

        # Apply header styles
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Apply row styles
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            worksheet.cell(row=row_idx, column=3).alignment = wrapping_alignment
            worksheet.cell(row=row_idx, column=5).alignment = center_and_wrap
            worksheet.cell(row=row_idx, column=6).alignment = wrapping_alignment

            if row_idx % 2 == 1:
                for cell in row[2:]:
                    cell.fill = even_row_fill
            
            covered_cell = row[3]
            covered_cell.alignment = centered_alignment
            if covered_cell.value == "✓":
                covered_cell.font = green_font
            elif covered_cell.value == "×":
                covered_cell.font = red_font

        # Set column widths
        column_widths = {'A': 10, 'B': 40, 'C': 30, 'D': 10, 'E': 30, 'F': 60}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Merge "Article" column if values are the same
        current_article = None
        start_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            article_cell = row[0]
            if article_cell.value != current_article:
                if start_row is not None:
                    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=row_idx - 1, end_column=1)
                    worksheet.merge_cells(start_row=start_row, start_column=2, end_row=row_idx - 1, end_column=2)
                    worksheet.cell(row=start_row, column=1).alignment = centered_alignment
                    worksheet.cell(row=start_row, column=2).alignment = wrapping_alignment
                    for col_idx in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_idx - 1, column=col_idx).border = bottom_border
                current_article = article_cell.value
                start_row = row_idx
        
        # Merge last group if applicable
        if start_row is not None:
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=worksheet.max_row, end_column=1)
            worksheet.merge_cells(start_row=start_row, start_column=2, end_row=worksheet.max_row, end_column=2)
            worksheet.cell(row=start_row, column=1).alignment = centered_alignment
            worksheet.cell(row=start_row, column=2).alignment = wrapping_alignment
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.cell(row=worksheet.max_row, column=col_idx).border = bottom_border

        worksheet.column_dimensions.group("B", "B", hidden=True)

    # Save the workbook to the buffer
    output.seek(0)
    return output
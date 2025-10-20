from docx import Document as dx
import pandas as pd
import re
import os
import ast
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from langchain_chroma import Chroma
from modules.model.open_source_llm import create_embeddings, create_openai_llm, create_anthropic_llm
from modules.prompts.gap_finder_prompt import build_gap_prompt, build_table_prompt
from langchain.schema.document import Document

# Initialize open source components
embeddings = create_embeddings()
# You can choose between OpenAI or Anthropic
llm = create_openai_llm()  # or create_anthropic_llm()

def parse_embedding(embedding_str):
    """
    Parse embedding string to numpy array
    """
    try:
        return np.array(ast.literal_eval(embedding_str))
    except Exception as e:
        print(f"Error parsing embedding: {embedding_str}, error: {e}")
        return None

def split_docx_by_bold_titles(file_path):
    """
    Split DOCX document by bold titles
    """
    doc = dx(file_path)
    
    content = []
    current_title = None
    index = 1
    main_title = None

    for paragraph in doc.paragraphs:
        if paragraph.style.name == 'List Paragraph':
            main_title = str(index) + '. ' + paragraph.text.strip()
            index += 1
            content.append([main_title, "", []])
            current_title = None
        elif paragraph.style.name != 'List Paragraph' and any(run.bold for run in paragraph.runs):
            current_title = paragraph.text.strip()
            content.append([main_title, current_title, []])
        elif main_title:
            text = paragraph.text.strip()
            if text:
                if current_title:
                    content[-1][2].append(text)
                else:
                    content[-1][2].append(text)

    return content

def create_vectorstore(db_directory, split_content):
    """
    Create vector store using open source embeddings
    """
    if not os.path.isdir(db_directory):
        to_embed_docs = []
        doc_num = 1
        
        for row in split_content:
            if row[2]:
                joined_par = '\n'.join(row[2])
                if row[1]:
                    content = f"Title: {row[0]}\n SubTitle: {row[1]}\n" + joined_par
                else:
                    content = f"Title: {row[0]}\n" + joined_par
                    
                print(f"Processing document {doc_num}")
                to_embed_docs.append(Document(
                    page_content=content,
                    metadata={
                        'id': doc_num,
                        'title': row[0],
                        'sub_title': row[1]
                    }
                ))
            doc_num += 1
            
        os.makedirs(db_directory)
        vectorstore = Chroma(
            collection_name="collection_document",
            embedding_function=embeddings,
            persist_directory=db_directory,
        )
        vectorstore.add_documents(to_embed_docs)
    else: 
        vectorstore = Chroma(
            collection_name="collection_document",
            embedding_function=embeddings,
            persist_directory=db_directory,
        )
    
    return vectorstore 

def extract_table_from_text(Article, Original_article, text):
    """
    Extract table from text response
    """
    lines = text.split("\n")
    rows = [line.strip("|").split("|") for line in lines if "|" in line and not re.match(r"^\|-+", line)]

    if len(rows) < 2:
        data = [[f"Rz.{Article}", Original_article, '', '', '', 'The provided string does not contain a valid table.']]
    else:
        data = [[f"Rz.{Article}"] + [Original_article] + [cell.replace('<br>', '\n').strip() for cell in row] for row in rows[1:]]
    
    for row in data:
        if row[3] == "Yes":
            row[3] = "✓"
        elif row[3] == "No":
            row[3] = "×"

    return data

def write_to_excel(df_table):
    """
    Write DataFrame to Excel with formatting
    """
    with pd.ExcelWriter('Results/Report_2017_formated_selected_open_source.xlsx', engine='openpyxl') as writer:
        df_table.to_excel(writer, index=False, sheet_name="Sheet1")

        workbook = writer.book
        worksheet = workbook['Sheet1']

        # Define styles
        header_fill = PatternFill(start_color="434fc3", end_color="434fc3", fill_type="solid")
        even_row_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        green_font = Font(color="008000", bold=True, size=12) 
        red_font = Font(color="fe4017", bold=True, size=14) 
        centered_alignment = Alignment(horizontal='center', vertical='center')
        wrapping_alignment = Alignment(wrap_text=True, vertical='top') 
        bottom_border = Border(bottom=Side(style="thin"))

        # Apply header styles
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Apply row styles
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            worksheet.cell(row=row_idx, column=3).alignment = wrapping_alignment
            worksheet.cell(row=row_idx, column=5).alignment = wrapping_alignment
            worksheet.cell(row=row_idx, column=6).alignment = wrapping_alignment
        
            if row_idx % 2 == 1:
                for cell in row[2:]:
                    cell.fill = even_row_fill
            
            covered_cell = row[3]
            covered_cell.alignment = centered_alignment
            if covered_cell.value == "✓":
                covered_cell.font = green_font                
            elif covered_cell.value == "×":
                covered_cell.font = red_font

        # Set column widths
        column_widths = {
            'A': 10, 'B': 40, 'C': 30, 'D': 10, 'E': 30, 'F': 60
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width 

        # Merge cells
        current_article = None
        start_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            article_cell = row[0]
            if article_cell.value != current_article:
                if start_row is not None:
                    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=row_idx - 1, end_column=1)
                    worksheet.merge_cells(start_row=start_row, start_column=2, end_row=row_idx - 1, end_column=2)
                    worksheet.cell(row=start_row, column=1).alignment = centered_alignment
                    worksheet.cell(row=start_row, column=2).alignment = wrapping_alignment
                    for col_idx in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_idx - 1, column=col_idx).border = bottom_border
                current_article = article_cell.value
                start_row = row_idx
                
        if start_row is not None:
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=worksheet.max_row, end_column=1)
            worksheet.merge_cells(start_row=start_row, start_column=2, end_row=worksheet.max_row, end_column=2)
            worksheet.cell(row=start_row, column=1).alignment = centered_alignment 
            worksheet.cell(row=start_row, column=2).alignment = wrapping_alignment
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.cell(row=worksheet.max_row, column=col_idx).border = bottom_border

        worksheet.column_dimensions.group('B', 'B', hidden=True) 

def run_gap_analysis():
    """
    Run the complete gap analysis using open source components
    """
    print("Starting gap analysis with open source components...")
    
    # Process document
    file_path = "Data/Document/Concept_Risk and Governance_EN.docx"  
    split_content = split_docx_by_bold_titles(file_path)

    # Create vector store
    db_directory_document = f"vectorestores/chroma_db_document_open_source"
    vectorstore = create_vectorstore(db_directory_document, split_content)

    # Load and process FINMA data
    df_2017 = pd.read_excel('Data/Finma_EN/splitted/finma2017.xlsx') 
    df_2017['Embedding'] = df_2017['Embedding'].apply(parse_embedding)
    
    results_df_2017 = pd.DataFrame(columns=['Margin', 'Finma2017', "Reports", "Table"])
    headers = ['Article', 'Article Content', 'Requirement', 'Covered', 'Reference in Document', 'Comment']
    all_rows = []

    # Process subset of data
    subset_df = df_2017.iloc[9:66]
    
    for i, row in subset_df.iterrows():
        print(f"Processing row {i}")
        
        full_text = ''
        title = str(row.get('Title', None))
        sub_title = str(row.get('SubTitle', None))
        sub_subtitle = str(row.get('Sub_Subtitle', None))
        margin = str(row.get('Margin', None))
        text = row.get('Text', None)
        embedded_item = row.get('Embedding', None)

        if text != 'Abrogated' and embedded_item is not None:
            # Find similar documents
            results = vectorstore.similarity_search_by_vector(
                embedding=embedded_item.tolist(), k=4
            )

            # Build full text
            if title and title != 'None' and title != 'nan':
                full_text = title
            if sub_title and sub_title != 'None' and sub_title != 'nan':
                full_text = full_text + '\n' + sub_title
            if sub_subtitle and sub_subtitle != 'None' and sub_subtitle != 'nan':
                full_text = full_text + '\n' + sub_subtitle
            if full_text:
                full_text = full_text + '\n' + text
            else:
                full_text = text

            # Generate gap analysis
            gap_prompt = build_gap_prompt(results, full_text)
            response = llm.ask_llm(gap_prompt, temperature=0.6)

            # Generate table
            table_prompt = build_table_prompt(response)   
            table_response = llm.ask_llm(table_prompt, temperature=0.1) 

            # Store results
            new_row = [{
                "Margin": margin,
                "Finma2017": full_text,
                "Reports": response,
                "Table": table_response
            }]
            results_df_2017 = pd.concat([results_df_2017, pd.DataFrame(new_row)], ignore_index=True)

            # Extract table data
            cur_table_data = extract_table_from_text(margin, full_text, table_response)
            all_rows = all_rows + cur_table_data

    # Save results
    results_df_2017.to_excel('Results/Report_2017_0.6_open_source.xlsx', index=False) 
    df_table = pd.DataFrame(all_rows, columns=headers)  
    write_to_excel(df_table)
    
    print("Gap analysis completed successfully!")

if __name__ == "__main__":
    run_gap_analysis()
